#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_unterrichtsplan.py
============================
Generates a structured Unterrichtsplan Excel file from a Zeitplan and Modulplan.

Requirements: pip install openpyxl pandas

Usage:
    python generate_unterrichtsplan.py \
        --zeitplan "Zeitplan_HH.xlsx" \
        --modulplan "Modulplan_KBM.xlsx" \
        --kursname "KBM" \
        --output "Unterrichtsplan_KBM_2026.xlsx" \
        --ue-pro-tag 9
"""

import argparse
import math
import os
import sys
from datetime import date, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"
ALT_ROW_BG = "DDEEFF"
META_BG = "F2F2F2"
SUMME_BG = "D9D9D9"
THIN_BORDER_COLOR = "BFBFBF"

# Column widths for Sheet 1
COL_WIDTHS_S1 = {"A": 10, "B": 46, "C": 13, "D": 13, "E": 8, "F": 8, "G": 32}

# Special day-type colours for Sheet 3 (calendar markings, NOT modules).
# (bg_hex, fg_hex) — the SAME hex is used for the CF rules and the legend swatches,
# so the calendar and its legend can never drift apart.
SPECIAL_DAY_COLORS: Dict[str, Tuple[str, str]] = {
    "Fe":   ("70AD47", "000000"),   # medium green  (Ferien)
    "Pr":   ("FFD966", "000000"),   # yellow        (Praktikum)
    "Pr+":  ("FFD966", "000000"),   # same yellow   (Praktikum + Fachpraktische Begleitung)
    "Prüf": ("FF7070", "FFFFFF"),   # red           (Prüfung)
    "Ft":   ("D5EDB3", "000000"),   # light green   (Feiertag, heller als Fe)
}

# Colour palette assigned IN ORDER to whatever module IDs actually occur.
# This lets the script handle a variable number of modules with arbitrary IDs
# (M1..M18, M0..M16, IM0..IM3, …) without any hardcoded module list.
MODULE_PALETTE: List[Tuple[str, str]] = [
    ("E2EEF9", "000000"), ("BDD7EE", "000000"), ("9DC3E6", "000000"),
    ("70ADD4", "FFFFFF"), ("D9EEF3", "000000"), ("A8D5DF", "000000"),
    ("6BB8C8", "FFFFFF"), ("3A9DB3", "FFFFFF"), ("E8E8F7", "000000"),
    ("CACAED", "000000"), ("A5A5DE", "000000"), ("7B7BCF", "FFFFFF"),
    ("D8E8F8", "000000"), ("B0CFED", "000000"), ("7FAFD9", "000000"),
    ("4D8EC4", "FFFFFF"), ("FCE4C8", "000000"), ("F9C99A", "000000"),
    ("F4A563", "000000"), ("EC8330", "FFFFFF"), ("BFBFE0", "000000"),
    ("C9E2B6", "000000"), ("9FD18B", "000000"), ("E0C7E8", "000000"),
]


def assign_module_colors(module_ids: List[str]) -> Dict[str, Tuple[str, str]]:
    """Assign a distinct palette colour to each module ID, in first-seen order.
    Cycles the palette if there are more modules than colours."""
    colors: Dict[str, Tuple[str, str]] = {}
    idx = 0
    for mid in module_ids:
        if mid in colors:
            continue
        colors[mid] = MODULE_PALETTE[idx % len(MODULE_PALETTE)]
        idx += 1
    return colors

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _thin_border() -> Border:
    side = Side(style="thin", color=HEADER_BG)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_fill(hex_color: str = HEADER_BG) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _make_font(bold: bool = False, size: int = 10, color: str = "000000",
               italic: bool = False) -> Font:
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)


def _excel_serial_to_date(serial: float) -> Optional[date]:
    """Convert an Excel date serial number to a Python date."""
    try:
        serial = int(serial)
        # Excel incorrectly treats 1900 as a leap year; offset accordingly
        if serial >= 60:
            serial -= 1
        delta = timedelta(days=serial - 1)
        return date(1900, 1, 1) + delta
    except (ValueError, TypeError, OverflowError):
        return None


def _normalize_col_name(name: str) -> str:
    if not isinstance(name, str):
        return ""
    return name.strip().lower()


# ---------------------------------------------------------------------------
# Zeitplan parsing
# ---------------------------------------------------------------------------

def _detect_year_sheets(wb: openpyxl.Workbook) -> List[str]:
    """Return sheet names that look like annual calendar sheets, in order."""
    annual_patterns = [
        "1. jahr", "2. jahr", "3. jahr", "4. jahr",
        "1.jahr", "2.jahr", "3.jahr", "4.jahr",
    ]
    result: List[str] = []
    # First try canonical year names
    for name in wb.sheetnames:
        if name.strip().lower() in annual_patterns:
            result.append(name)
    if result:
        return result
    # Fall back: sheets whose name is a 4-digit year
    for name in wb.sheetnames:
        if name.strip().isdigit() and 2000 <= int(name.strip()) <= 2100:
            result.append(name)
    return result



def _parse_annual_sheet(ws) -> List[Tuple[date, Optional[str]]]:
    """
    Parse one annual-calendar worksheet.
    Returns (date, day_type) tuples. day_type=None means teaching day.
    Relies solely on cell values — the Zeitplan is pre-marked.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Find month header row: row with >= 3 date/datetime objects
    month_row_idx = None
    month_col_map: Dict[int, date] = {}

    for row_idx, row in enumerate(rows[:10]):
        candidates: Dict[int, date] = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            if hasattr(cell, "year"):  # datetime or date object
                candidates[col_idx] = date(cell.year, cell.month, 1)
            elif isinstance(cell, (int, float)) and 10000 < cell < 60000:
                d = _excel_serial_to_date(cell)
                if d:
                    candidates[col_idx] = date(d.year, d.month, 1)
        if len(candidates) >= 3:
            month_row_idx = row_idx
            month_col_map = candidates
            break

    if month_row_idx is None or not month_col_map:
        print(f"  [WARN] Sheet '{ws.title}': no month header row found. Skipping.")
        return []

    # Two-column-per-month layout: month header at col N, actual content at col N+1
    # Detect by average gap between header columns
    sorted_header_cols = sorted(month_col_map.keys())
    if len(sorted_header_cols) >= 2:
        gaps = [sorted_header_cols[i+1] - sorted_header_cols[i]
                for i in range(len(sorted_header_cols) - 1)]
        if sum(gaps) / len(gaps) >= 1.8:
            # Content column is one to the right of the month header column
            month_col_map = {col_idx + 1: v for col_idx, v in month_col_map.items()}

    # Find day-number column: first column before month columns with integers 1-31
    day_col_idx = None
    min_month_col = min(month_col_map.keys())
    for col_idx in range(min_month_col):
        col_vals = [rows[r][col_idx] if col_idx < len(rows[r]) else None
                    for r in range(month_row_idx + 1, min(month_row_idx + 35, len(rows)))]
        int_vals = [v for v in col_vals if isinstance(v, (int, float)) and 1 <= v <= 31]
        if len(int_vals) >= 20:
            day_col_idx = col_idx
            break
    if day_col_idx is None:
        day_col_idx = max(0, min_month_col - 1)

    # Build calendar entries
    calendar: List[Tuple[date, Optional[str]]] = []
    sorted_month_cols = sorted(month_col_map.keys())

    for row_idx in range(month_row_idx + 1, len(rows)):
        row = rows[row_idx]
        if not row:
            continue
        day_cell = row[day_col_idx] if day_col_idx < len(row) else None
        if day_cell is None:
            continue
        try:
            day_num = int(day_cell)
        except (ValueError, TypeError):
            continue
        if not (1 <= day_num <= 31):
            continue

        for col_idx in sorted_month_cols:
            month_first = month_col_map[col_idx]
            try:
                current_date = date(month_first.year, month_first.month, day_num)
            except ValueError:
                continue  # invalid date (e.g. Feb 30)

            cell_val = row[col_idx] if col_idx < len(row) else None
            day_type = _classify_day(current_date, cell_val)
            calendar.append((current_date, day_type))

    return calendar


def _classify_day(d: date, cell_val) -> Optional[str]:
    """
    Classify a calendar day based solely on the Zeitplan cell value.
    The Zeitplan is the single source of truth:
      - numeric value (int/float) = Unterrichtstag (teaching day)
      - 'Fe'  = Ferien
      - 'Pr'  = Praktikum
      - 'Pr/4'= Fachpraktische Begleitung (FPB) -> mapped to 'Pr+'
      - 'Prüf'= Prüfungstag
      - 'Ft'  = Feiertag (already marked in Zeitplan)
      - None/empty = Wochenende or non-course day
    """
    if isinstance(cell_val, (int, float)):
        return None  # numeric UE value = teaching day
    if isinstance(cell_val, str):
        val = cell_val.strip()
        if not val:
            return "WE"
        mapping = {
            "fe":       "Fe",
            "ferien":   "Fe",
            "pr":       "Pr",
            "praktikum":"Pr",
            "pr/4":     "Pr+",
            "prüf":     "Prüf",
            "prüfung":  "Prüf",
            "pruef":    "Prüf",
            "ft":       "Ft",
            "feiertag": "Ft",
        }
        return mapping.get(val.lower(), val)
    # None/empty cell on a weekday = Feiertag (grey fill, no text/UE value in Zeitplan)
    # None/empty on Sat/Sun = Wochenende
    if d.weekday() >= 5:
        return "WE"
    return "Ft"


def parse_zeitplan(path: str) -> List[Tuple[date, Optional[str]]]:
    """
    Parse the Zeitplan Excel file.
    The Zeitplan is the single source of truth for all day types.
    Returns ordered list of (date, day_type). Teaching days have day_type=None.
    """
    print(f"[INFO] Parsing Zeitplan: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)

    annual_sheets = _detect_year_sheets(wb)
    if not annual_sheets:
        print("  [WARN] No annual sheets detected. Trying all non-special sheets.")
        skip = {"schulfreie", "feiertag", "legende", "hinweise", "info", "ubersicht", "ubersicht"}
        annual_sheets = [
            s for s in wb.sheetnames
            if not any(kw in s.lower() for kw in skip)
        ]

    all_days: List[Tuple[date, Optional[str]]] = []
    seen_dates: set = set()

    for sheet_name in annual_sheets:
        print(f"  Processing sheet: '{sheet_name}'")
        ws = wb[sheet_name]
        sheet_days = _parse_annual_sheet(ws)
        added = 0
        for entry in sheet_days:
            if entry[0] not in seen_dates:
                all_days.append(entry)
                seen_dates.add(entry[0])
                added += 1
        print(f"    -> {added} calendar days added.")

    all_days.sort(key=lambda x: x[0])

    # Detect course Beginn/Ende from first sheet header rows
    course_start: Optional[date] = None
    course_end:   Optional[date] = None
    if annual_sheets:
        header_ws = wb[annual_sheets[0]]
        for row in list(header_ws.iter_rows(values_only=True))[:6]:
            if not row:
                continue
            for ci, cell in enumerate(row):
                if not isinstance(cell, str):
                    continue
                if "beginn" in cell.lower():
                    for adj in row[ci+1:ci+4]:
                        if hasattr(adj, "year"):
                            course_start = adj.date() if hasattr(adj, "date") else adj
                            break
                if "ende" in cell.lower():
                    for adj in row[ci+1:ci+4]:
                        if hasattr(adj, "year"):
                            course_end = adj.date() if hasattr(adj, "date") else adj
                            break

    if course_start:
        print(f"  Course start: {course_start} -> filtering.")
        all_days = [(d, dt) for d, dt in all_days if d >= course_start]
    if course_end:
        print(f"  Course end:   {course_end}")
        all_days = [(d, dt) for d, dt in all_days if d <= course_end]

    teaching_count = sum(1 for _, dt in all_days if dt is None)
    pr_count = sum(1 for _, dt in all_days if dt == "Pr")
    pr_plus_count = sum(1 for _, dt in all_days if dt == "Pr+")
    fe_count = sum(1 for _, dt in all_days if dt == "Fe")
    ft_count = sum(1 for _, dt in all_days if dt == "Ft")
    print(f"  Total: {len(all_days)} days | Teaching: {teaching_count} | "
          f"Fe: {fe_count} | Pr: {pr_count} | Pr+: {pr_plus_count} | Ft: {ft_count}")
    return all_days


# ---------------------------------------------------------------------------
# Modulplan parsing
# ---------------------------------------------------------------------------

def parse_modulplan(path: str) -> List[Dict]:
    """
    Parse the Modulplan Excel.
    Returns list of dicts: {id, title, ue}
    """
    print(f"[INFO] Parsing Modulplan: {path}")
    df = pd.read_excel(path, header=None, dtype=str)

    # Find header row (first row containing at least 2 recognizable column names)
    header_row_idx = None
    col_map: Dict[str, int] = {}

    id_aliases = {"modul", "nr", "modul-nr", "modul nr", "modul_nr"}
    title_aliases = {"modulbezeichnung", "bezeichnung", "name", "titel", "modul-bezeichnung"}
    ue_aliases = {"ue", "unterrichtseinheiten", "stunden", "u.e.", "ue/std"}

    for row_idx, row in df.iterrows():
        row_lower = {_normalize_col_name(str(v)): ci for ci, v in enumerate(row) if pd.notna(v)}
        found_id = next((ci for k, ci in row_lower.items() if k in id_aliases), None)
        found_title = next((ci for k, ci in row_lower.items() if k in title_aliases), None)
        found_ue = next((ci for k, ci in row_lower.items() if k in ue_aliases), None)
        if found_id is not None and found_title is not None and found_ue is not None:
            header_row_idx = row_idx
            col_map = {"id": found_id, "title": found_title, "ue": found_ue}
            break
        elif found_title is not None and found_ue is not None:
            header_row_idx = row_idx
            col_map = {"id": found_id, "title": found_title, "ue": found_ue}
            break

    if header_row_idx is None:
        raise ValueError(
            "Modulplan: Could not find a header row with recognizable columns. "
            "Expected: Modul/Nr, Modulbezeichnung/Bezeichnung, UE/Stunden."
        )

    modules: List[Dict] = []
    for row_idx in range(header_row_idx + 1, len(df)):
        row = df.iloc[row_idx]

        def _get(ci):
            return row.iloc[ci] if ci is not None and ci < len(row) else None

        raw_id = _get(col_map.get("id"))
        raw_title = _get(col_map["title"])
        raw_ue = _get(col_map["ue"])

        # Skip empty rows
        if pd.isna(raw_title) or str(raw_title).strip() == "":
            continue
        if pd.isna(raw_ue) or str(raw_ue).strip() in ("", "nan"):
            continue

        try:
            ue_val = float(str(raw_ue).strip().replace(",", "."))
        except ValueError:
            continue

        if ue_val <= 0:
            continue

        mod_id = str(raw_id).strip() if raw_id is not None and not pd.isna(raw_id) else f"M{len(modules)}"

        modules.append({
            "id": mod_id,
            "title": str(raw_title).strip(),
            "ue": ue_val,
        })

    print(f"  Found {len(modules)} modules, total UE: {sum(m['ue'] for m in modules):.0f}")
    return modules


# ---------------------------------------------------------------------------
# Scheduling
# ---------------------------------------------------------------------------

def check_deficit(
    teaching_days: List[date], modules: List[Dict], ue_pro_tag: int
) -> int:
    """Return surplus (positive) or deficit (negative) in teaching days."""
    total_days_needed = sum(math.ceil(m["ue"] / ue_pro_tag) for m in modules)
    return len(teaching_days) - total_days_needed


def _is_fpb_module(module: Dict) -> bool:
    """True for the Fachpraktische Begleitung (FPB) — must NEVER be shortened."""
    text = f"{module.get('id', '')} {module.get('title', '')}".lower()
    norm = text.replace("-", "").replace(" ", "")
    return (
        "fachpraktischebegleitung" in norm
        or module.get("id", "").strip().lower() == "fpb"
        or "(fpb)" in text
    )


def _is_pv_module(module: Dict) -> bool:
    """True for a Prüfungsvorbereitung (exam-prep buffer module).
    The FPB is explicitly excluded so it can never be mistaken for the buffer."""
    if _is_fpb_module(module):
        return False
    text = f"{module.get('id', '')} {module.get('title', '')}".lower()
    norm = text.replace("ü", "ue").replace("-", "").replace(" ", "")
    keywords = ("pruefungsvorbereitung", "pruefungsvorb", "prfungsvorbereitung")
    if any(k in norm for k in keywords):
        return True
    return module.get("id", "").strip().lower() in ("pv", "pvb")


def auto_adjust_pruefungsvorbereitung(
    modules: List[Dict], available_days: int, ue_pro_tag: int
) -> Tuple[List[Dict], Optional[List[Dict]]]:
    """
    Resolve a teaching-day deficit by shortening ONLY the Prüfungsvorbereitung
    module(s). Rationale: the Modulplan is idealtypisch (carries maximum UE);
    when the concrete Zeitplan offers fewer teaching days, the PV buffer absorbs
    the difference. The Fachpraktische Begleitung is never shortened.

    Returns (modules, adjustments):
      - adjustments = [] when no change was needed,
      - adjustments = list of {id, old_ue, new_ue, days_removed} on success,
      - adjustments = None when a deficit exists but no PV module was found.
    """
    total_needed = sum(math.ceil(m["ue"] / ue_pro_tag) for m in modules)
    deficit = total_needed - available_days
    if deficit <= 0:
        return modules, []

    pv_indices = [i for i, m in enumerate(modules) if _is_pv_module(m)]
    if not pv_indices:
        return modules, None  # signal: nothing to absorb the deficit

    adjustments: List[Dict] = []
    remaining = deficit
    # Shorten from the LAST Prüfungsvorbereitung block backwards.
    for i in reversed(pv_indices):
        if remaining <= 0:
            break
        m = modules[i]
        cur_days = math.ceil(m["ue"] / ue_pro_tag)
        removable = max(0, cur_days - 1)  # keep at least one day
        take = min(removable, remaining)
        if take <= 0:
            continue
        new_days = cur_days - take
        old_ue = m["ue"]
        new_ue = new_days * ue_pro_tag
        m["ue"] = new_ue
        m["bemerkung"] = (
            f"Prüfungsvorbereitung gekürzt von {int(old_ue)} auf {int(new_ue)} UE "
            f"(Anpassung an verfügbare Unterrichtstage)"
        )
        adjustments.append({
            "id": m["id"], "old_ue": old_ue, "new_ue": new_ue, "days_removed": take,
        })
        remaining -= take

    return modules, adjustments


def schedule_modules(
    calendar_days: List[Tuple[date, Optional[str]]],
    modules: List[Dict],
    ue_pro_tag: int,
) -> Tuple[List[Dict], Dict[date, str]]:
    """
    Assign teaching days to modules.

    Returns:
        schedule: list of dicts with id, title, ue, start_date, end_date,
                  days_assigned, ue_assigned
        day_to_module: dict mapping each calendar date to its label
                       (module ID, 'Fe', 'Pr', 'Prüf', 'Ft', 'WE', or '')
    """
    teaching_days = [d for d, dt in calendar_days if dt is None]
    day_to_type: Dict[date, Optional[str]] = {d: dt for d, dt in calendar_days}

    teaching_idx = 0
    schedule: List[Dict] = []
    day_to_module: Dict[date, str] = {}

    # Fill non-teaching days first
    for d, dt in calendar_days:
        if dt == "WE":
            day_to_module[d] = ""
        elif dt is not None:
            day_to_module[d] = dt  # Fe, Pr, Pr+, Prüf, Ft

    for module in modules:
        days_needed = math.ceil(module["ue"] / ue_pro_tag)
        assigned_teaching_days: List[date] = []

        for _ in range(days_needed):
            if teaching_idx >= len(teaching_days):
                break
            day = teaching_days[teaching_idx]
            assigned_teaching_days.append(day)
            day_to_module[day] = module["id"]
            teaching_idx += 1

        if not assigned_teaching_days:
            print(f"  [WARN] Module '{module['id']}' has no teaching days assigned (ran out).")
            schedule.append({
                "id": module["id"],
                "title": module["title"],
                "ue": module["ue"],
                "start_date": None,
                "end_date": None,
                "days_assigned": 0,
                "ue_assigned": 0,
                "bemerkung": module.get("bemerkung", ""),
            })
            continue

        start_date = assigned_teaching_days[0]
        end_date = assigned_teaching_days[-1]
        days_assigned = len(assigned_teaching_days)
        ue_assigned = min(module["ue"], days_assigned * ue_pro_tag)

        schedule.append({
            "id": module["id"],
            "title": module["title"],
            "ue": module["ue"],
            "start_date": start_date,
            "end_date": end_date,
            "days_assigned": days_assigned,
            "ue_assigned": ue_assigned,
            "bemerkung": module.get("bemerkung", ""),
        })

    return schedule, day_to_module


# ---------------------------------------------------------------------------
# Excel building helpers
# ---------------------------------------------------------------------------

def _apply_header_row(ws, row: int, headers: List[str], col_start: int = 1):
    """Write a header row with standard blue styling."""
    for ci, header in enumerate(headers, start=col_start):
        cell = ws.cell(row=row, column=ci, value=header)
        cell.font = _make_font(bold=True, color=HEADER_FG, size=10)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()


def _write_s1_unterrichtsplan(
    ws, schedule: List[Dict], kursname: str, calendar_days: List[Tuple[date, Optional[str]]]
):
    """Build Sheet 1: Unterrichtsplan."""
    # Row 1: Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Unterrichtsplan – {kursname}"
    title_cell.font = Font(name="Arial", bold=True, size=16, color=HEADER_FG)
    title_cell.fill = _header_fill()
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2: Metadata strip
    from datetime import date as date_cls
    total_ue = sum(e["ue"] for e in schedule)
    total_days = sum(e["days_assigned"] for e in schedule)
    start_d = min((e["start_date"] for e in schedule if e["start_date"]), default=None)
    end_d = max((e["end_date"] for e in schedule if e["end_date"]), default=None)
    duration_str = ""
    if start_d and end_d:
        duration_str = f"{start_d.strftime('%d.%m.%Y')} – {end_d.strftime('%d.%m.%Y')}"

    meta_fill = PatternFill("solid", fgColor=META_BG.replace("#", ""))
    meta_font = _make_font(italic=True, size=10)
    meta_cells = [
        f"Erstellt: {date_cls.today().strftime('%d.%m.%Y')}",
        f"Gesamtdauer: {duration_str}",
        f"Gesamt-UE: {total_ue:.0f}",
        f"Gesamt-Tage: {total_days}",
        "", "", "",
    ]
    ws.merge_cells("A2:G2")
    ws["A2"].value = "  |  ".join(c for c in meta_cells if c)
    ws["A2"].font = meta_font
    ws["A2"].fill = meta_fill
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # Row 3: Spacer
    ws.row_dimensions[3].height = 6

    # Row 4: Headers
    headers = ["Modul", "Modulbezeichnung", "Beginn", "Ende", "Tage", "UE", "Bemerkungen"]
    _apply_header_row(ws, row=4, headers=headers)
    ws.row_dimensions[4].height = 18

    # Rows 5+: Data
    sum_ue = 0
    sum_days = 0
    for row_offset, entry in enumerate(schedule):
        row = 5 + row_offset
        bg = "FFFFFF" if row_offset % 2 == 0 else ALT_ROW_BG
        fill = PatternFill("solid", fgColor=bg)
        font = _make_font(size=10)

        def _write(col, val, num_format=None, align="left"):
            c = ws.cell(row=row, column=col, value=val)
            c.font = font
            c.fill = fill
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = _thin_border()
            if num_format:
                c.number_format = num_format
            return c

        _write(1, entry["id"], align="center")
        _write(2, entry["title"])
        if entry["start_date"]:
            _write(3, entry["start_date"], num_format="DD.MM.YYYY", align="center")
            _write(4, entry["end_date"], num_format="DD.MM.YYYY", align="center")
        else:
            _write(3, "—", align="center")
            _write(4, "—", align="center")
        _write(5, entry["days_assigned"], align="center")
        _write(6, int(entry["ue"]), align="center")
        _write(7, entry.get("bemerkung", ""))
        ws.row_dimensions[row].height = 15

        sum_ue += entry["ue"]
        sum_days += entry["days_assigned"]

    # SUMME row
    summe_row = 5 + len(schedule)
    summe_fill = PatternFill("solid", fgColor=SUMME_BG)
    summe_font = _make_font(bold=True, size=10)
    for col in range(1, 8):
        c = ws.cell(row=summe_row, column=col)
        c.fill = summe_fill
        c.font = summe_font
        c.border = _thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=summe_row, column=1).value = "SUMME"
    ws.cell(row=summe_row, column=5).value = sum_days
    ws.cell(row=summe_row, column=6).value = int(sum_ue)
    ws.row_dimensions[summe_row].height = 16

    # Column widths
    for col_letter, width in COL_WIDTHS_S1.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A5"


def _write_s2_kalendar_uebersicht(
    ws, calendar_days: List[Tuple[date, Optional[str]]], day_to_module: Dict[date, str], kursname: str
):
    """Build Sheet 2: Kalenderübersicht."""
    # Title
    ws.merge_cells("A1:E1")
    ws["A1"].value = f"Kalenderübersicht – {kursname}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=HEADER_FG)
    ws["A1"].fill = _header_fill()
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Headers
    headers = ["Jahr", "Monat", "Unterrichtstage", "UE", "Kumuliert UE"]
    _apply_header_row(ws, row=2, headers=headers)
    ws.row_dimensions[2].height = 16

    # Group teaching days by year/month
    from collections import defaultdict
    monthly: Dict[Tuple[int, int], int] = defaultdict(int)
    ue_per_day_approx = 9  # will be overridden if we can access ue_pro_tag

    for d, dt in calendar_days:
        if dt is None:
            monthly[(d.year, d.month)] += 1

    cumulative_ue = 0
    month_names_de = [
        "", "Januar", "Februar", "März", "April", "Mai", "Juni",
        "Juli", "August", "September", "Oktober", "November", "Dezember"
    ]
    row = 3
    for (year, month), count in sorted(monthly.items()):
        ue = count * ue_per_day_approx
        cumulative_ue += ue
        bg = "FFFFFF" if (row % 2 == 0) else ALT_ROW_BG
        fill = PatternFill("solid", fgColor=bg)
        font = _make_font(size=10)

        values = [year, month_names_de[month], count, ue, cumulative_ue]
        for ci, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = font
            c.fill = fill
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="center" if ci != 2 else "left", vertical="center")
        ws.row_dimensions[row].height = 14
        row += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "A3"


def _write_s3_jahreskalender(
    ws,
    calendar_days: List[Tuple[date, Optional[str]]],
    day_to_module: Dict[date, str],
    schedule: List[Dict],
    kursname: str,
):
    """Build Sheet 3: Jahreskalender with conditional formatting."""
    from collections import defaultdict

    # Organise calendar_days by (year, month) -> {day: label}
    by_month: Dict[Tuple[int, int], Dict[int, str]] = defaultdict(dict)
    for d, dt in calendar_days:
        label = day_to_module.get(d, "")
        if label is None:
            label = ""
        by_month[(d.year, d.month)][d.day] = label

    # Determine display range from the module schedule (not all calendar_days)
    sched_dates = [e["start_date"] for e in schedule if e["start_date"]] + \
                  [e["end_date"]   for e in schedule if e["end_date"]]
    if sched_dates:
        course_start = min(sched_dates)
        course_end   = max(sched_dates)
    else:
        all_data = sorted(by_month.keys())
        course_start = date(all_data[0][0],  all_data[0][1],  1)
        course_end   = date(all_data[-1][0], all_data[-1][1], 28)

    # Border month before/after the course
    cs_y, cs_m = course_start.year, course_start.month
    ce_y, ce_m = course_end.year,   course_end.month
    border_before = (cs_y - 1, 12) if cs_m == 1 else (cs_y, cs_m - 1)
    border_after  = (ce_y + 1,  1) if ce_m == 12 else (ce_y, ce_m + 1)

    # Build sorted_months covering border_before … border_after
    b_year, b_month = border_before
    e_year, e_month = border_after
    sorted_months: List[Tuple[int, int]] = []
    y, m = b_year, b_month
    while (y, m) <= (e_year, e_month):
        sorted_months.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1

    if not sorted_months:
        ws["A1"].value = "Keine Daten."
        return

    # Group months by year for row 2 headers
    years: Dict[int, List[int]] = defaultdict(list)
    for year, month in sorted_months:
        years[year].append(month)

    # Layout constants
    month_col_start = 2  # Col A = day numbers, Col B = first month
    n_months = len(sorted_months)
    last_data_col = month_col_start + n_months - 1
    last_col_letter = get_column_letter(last_data_col)
    data_range = f"B4:{last_col_letter}34"

    # Shared styles
    grey_side = Side(style="thin", color="BFBFBF")
    grey_border = Border(left=grey_side, right=grey_side, top=grey_side, bottom=grey_side)
    weekend_fill = PatternFill("solid", fgColor="AEAAAA")
    invalid_fill = PatternFill("solid", fgColor="D9D9D9")
    cal_font = Font(name="Arial", size=9)

    # ---- Row 1: Title ----
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"].value = f"Jahreskalender – {kursname}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=HEADER_FG)
    ws["A1"].fill = _header_fill()
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ---- Row 2: Year group headers ----
    col_cursor = month_col_start
    for year, months_in_year in sorted(years.items()):
        span = len(months_in_year)
        sl = get_column_letter(col_cursor)
        el = get_column_letter(col_cursor + span - 1)
        if span > 1:
            ws.merge_cells(f"{sl}2:{el}2")
        c = ws.cell(row=2, column=col_cursor, value=str(year))
        c.font = Font(name="Arial", bold=True, size=10, color=HEADER_FG)
        c.fill = _header_fill()
        c.alignment = Alignment(horizontal="center", vertical="center")
        col_cursor += span
    ws["A2"].value = "Tag"
    ws["A2"].font = Font(name="Arial", bold=True, size=9, color=HEADER_FG)
    ws["A2"].fill = _header_fill()
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ---- Row 3: Month headers ----
    month_names_short = ["", "Jan", "Feb", "Mär", "Apr", "Mai", "Jun",
                         "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"]
    ws["A3"].fill = _header_fill()
    ws["A3"].font = Font(name="Arial", bold=True, size=9, color=HEADER_FG)
    ws.row_dimensions[3].height = 16

    for col_idx, (year, month) in enumerate(sorted_months, start=month_col_start):
        lbl = f"{month_names_short[month]} {str(year)[2:]}"
        c = ws.cell(row=3, column=col_idx, value=lbl)
        c.font = Font(name="Arial", bold=True, color=HEADER_FG, size=9)
        c.fill = _header_fill()
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ---- Rows 4-34: Day rows ----
    for day_num in range(1, 32):
        row = 3 + day_num
        ws.row_dimensions[row].height = 15

        a_cell = ws.cell(row=row, column=1, value=day_num)
        a_cell.font = Font(name="Arial", bold=True, size=9)
        a_cell.fill = PatternFill("solid", fgColor="E8E8E8")
        a_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, (year, month) in enumerate(sorted_months, start=month_col_start):
            c = ws.cell(row=row, column=col_idx)
            c.font = cal_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = grey_border

            try:
                d = date(year, month, day_num)
            except ValueError:
                c.fill = invalid_fill
                c.value = None
                continue

            if d.weekday() >= 5:
                c.fill = weekend_fill
                c.value = None
            else:
                cell_label = by_month[(year, month)].get(day_num, "")
                c.value = cell_label if cell_label else None
                # No static fill — conditional formatting handles colours

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 4
    for col_idx in range(month_col_start, month_col_start + n_months):
        ws.column_dimensions[get_column_letter(col_idx)].width = 6.5

    ws.freeze_panes = "B4"

    # ---- Build the colour map: dynamic module colours + fixed special types ----
    # Modules get palette colours in schedule order; specials are fixed.
    # This single map drives BOTH the conditional formatting AND the legend,
    # so the calendar and its legend always agree.
    module_colors = assign_module_colors([e["id"] for e in schedule])
    day_colors: Dict[str, Tuple[str, str]] = {}
    day_colors.update(module_colors)
    day_colors.update(SPECIAL_DAY_COLORS)  # specials win over any ID collision

    # ---- Conditional formatting — type="expression", exact equality ----
    # Use bgColor (correct for DifferentialStyle in Excel CF XML).
    # Add modules first (lowest priority), specials last (highest priority).
    specials = ["Fe", "Ft", "Pr", "Pr+", "Prüf"]
    module_labels = [lbl for lbl in day_colors if lbl not in specials]
    final_order = module_labels + [s for s in specials if s in day_colors]

    for lbl in final_order:
        bg_hex, fg_hex = day_colors[lbl]
        dxf = DifferentialStyle(
            fill=PatternFill(bgColor=bg_hex),
            font=Font(name="Arial", color=fg_hex, size=9),
        )
        rule = Rule(type="expression", dxf=dxf)
        rule.formula = [f'B4="{lbl}"']
        ws.conditional_formatting.add(data_range, rule)

    # ---- Legend (starting at row 37) ----
    LEGEND_START = 37

    ws.cell(row=LEGEND_START, column=1).value = "Legende"
    ws.cell(row=LEGEND_START, column=1).font = Font(name="Arial", bold=True, size=9)
    ws.row_dimensions[LEGEND_START].height = 14

    # Headers for Tage / UE
    hdr_font = Font(name="Arial", bold=True, size=9)
    hdr_align = Alignment(horizontal="center", vertical="center")
    ws.cell(row=LEGEND_START, column=11, value="Tage").font = hdr_font
    ws.cell(row=LEGEND_START, column=11).alignment = hdr_align
    ws.cell(row=LEGEND_START, column=12, value="UE").font = hdr_font
    ws.cell(row=LEGEND_START, column=12).alignment = hdr_align

    legend_row = LEGEND_START + 1

    def _add_legend_row(lbl, description, bg_hex, fg_hex, tage_formula=None, ue_val=None):
        nonlocal legend_row
        # Col A: coloured swatch + label text
        sw = ws.cell(row=legend_row, column=1, value=lbl)
        sw.fill = PatternFill("solid", fgColor=bg_hex)
        sw.font = Font(name="Arial", size=9, color=fg_hex, bold=True)
        sw.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[legend_row].height = 14

        # Cols B:J merged — description
        ws.merge_cells(f"B{legend_row}:J{legend_row}")
        dc = ws.cell(row=legend_row, column=2, value=description)
        dc.font = Font(name="Arial", size=9)
        dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Col K: Tage (COUNTIF formula)
        kc = ws.cell(row=legend_row, column=11)
        if tage_formula is not None:
            kc.value = tage_formula
        kc.font = Font(name="Arial", size=9)
        kc.alignment = Alignment(horizontal="center", vertical="center")

        # Col L: UE
        lc = ws.cell(row=legend_row, column=12)
        if ue_val is not None:
            lc.value = ue_val
        lc.font = Font(name="Arial", size=9)
        lc.alignment = Alignment(horizontal="center", vertical="center")

        legend_row += 1

    cf = f"$B$4:${last_col_letter}$34"   # COUNTIF range

    # Special-day rows — swatch colours come from the SAME map as the calendar.
    _add_legend_row("Fe",   "Ferien",            *SPECIAL_DAY_COLORS["Fe"],
                    f'=COUNTIF({cf},"Fe")')
    _add_legend_row("Pr",   "Praktikumstag",     *SPECIAL_DAY_COLORS["Pr"],
                    f'=COUNTIF({cf},"Pr")')

    # Pr+ row — description contains "zzgl. 2 UE" for the dynamic UE formula
    pr_star_row = legend_row
    pr_star_desc = "Praktikumstag zzgl. 2 UE Fachpraktische Begleitung"
    _add_legend_row("Pr+",  pr_star_desc,        *SPECIAL_DAY_COLORS["Pr+"],
                    f'=COUNTIF({cf},"Pr+")')
    # L formula reads "2" from the description text in B{pr_star_row}
    ws.cell(row=pr_star_row, column=12).value = (
        f'=VALUE(TRIM(MID(B{pr_star_row},FIND("zzgl.",B{pr_star_row})+6,'
        f'FIND("UE",B{pr_star_row})-FIND("zzgl.",B{pr_star_row})-6)))*K{pr_star_row}'
    )

    _add_legend_row("Prüf", "Prüfungstag",       *SPECIAL_DAY_COLORS["Prüf"],
                    f'=COUNTIF({cf},"Prüf")')
    _add_legend_row("Ft",   "Feiertag (gesetzlich)", *SPECIAL_DAY_COLORS["Ft"],
                    f'=COUNTIF({cf},"Ft")')

    # Module rows — iterate the ACTUAL schedule (variable count), in order.
    for e in schedule:
        lbl = e["id"]
        if lbl in SPECIAL_DAY_COLORS:
            continue
        bg_hex, fg_hex = day_colors.get(lbl, ("FFFFFF", "000000"))
        _add_legend_row(
            lbl, e["title"], bg_hex, fg_hex,
            f'=COUNTIF({cf},"{lbl}")',
            int(e["ue"]),
        )

    # SUMME row
    ws.cell(row=legend_row, column=1, value="Summe").font = Font(name="Arial", bold=True, size=9)
    ws.row_dimensions[legend_row].height = 14
    sum_k = ws.cell(row=legend_row, column=11)
    sum_k.value = f"=SUM(K{LEGEND_START + 1}:K{legend_row - 1})"
    sum_k.font = Font(name="Arial", bold=True, size=9)
    sum_k.alignment = Alignment(horizontal="center", vertical="center")
    sum_l = ws.cell(row=legend_row, column=12)
    sum_l.value = f"=SUM(L{LEGEND_START + 1}:L{legend_row - 1})"
    sum_l.font = Font(name="Arial", bold=True, size=9)
    sum_l.alignment = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Main build function
# ---------------------------------------------------------------------------

def build_excel(
    schedule: List[Dict],
    calendar_days: List[Tuple[date, Optional[str]]],
    day_to_module: Dict[date, str],
    kursname: str,
    output_path: str,
    ue_pro_tag: int,
):
    """Write the three-sheet output Excel file."""
    print(f"[INFO] Building Excel: {output_path}")
    wb = Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Unterrichtsplan"
    _write_s1_unterrichtsplan(ws1, schedule, kursname, calendar_days)

    # Sheet 2
    ws2 = wb.create_sheet("Kalenderübersicht")
    _write_s2_kalendar_uebersicht(ws2, calendar_days, day_to_module, kursname)

    # Sheet 3
    ws3 = wb.create_sheet("Jahreskalender")
    _write_s3_jahreskalender(ws3, calendar_days, day_to_module, schedule, kursname)

    wb.save(output_path)
    print(f"[INFO] Saved: {output_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generates a structured Unterrichtsplan Excel from Zeitplan and Modulplan.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--zeitplan", required=True, help="Path to Zeitplan Excel file")
    parser.add_argument("--modulplan", required=True, help="Path to Modulplan Excel file")
    parser.add_argument("--kursname", default="Kurs", help="Course name (default: Kurs)")
    parser.add_argument("--output", default=None, help="Output file path (auto-generated if omitted)")
    parser.add_argument("--ue-pro-tag", type=int, default=9, dest="ue_pro_tag",
                        help="UE per teaching day (default: 9)")
    args = parser.parse_args()

    # Validate inputs
    if not os.path.isfile(args.zeitplan):
        print(f"[ERROR] Zeitplan not found: {args.zeitplan}", file=sys.stderr)
        sys.exit(1)
    if not os.path.isfile(args.modulplan):
        print(f"[ERROR] Modulplan not found: {args.modulplan}", file=sys.stderr)
        sys.exit(1)

    # Determine output path
    if args.output:
        output_path = args.output
    else:
        zeitplan_dir = os.path.dirname(os.path.abspath(args.zeitplan))
        # Try to detect start year
        try:
            wb_tmp = openpyxl.load_workbook(args.zeitplan, data_only=True, read_only=True)
            sheets = _detect_year_sheets(wb_tmp)
            wb_tmp.close()
            # If sheets are named like "2026", use that; else use current year
            start_year = sheets[0] if sheets and sheets[0].isdigit() else str(date.today().year)
        except Exception:
            start_year = str(date.today().year)
        filename = f"Unterrichtsplan_{args.kursname}_{start_year}.xlsx"
        output_path = os.path.join(zeitplan_dir, filename)

    # Parse inputs
    calendar_days = parse_zeitplan(args.zeitplan)
    modules = parse_modulplan(args.modulplan)

    if not calendar_days:
        print("[ERROR] No calendar days parsed from Zeitplan.", file=sys.stderr)
        sys.exit(1)
    if not modules:
        print("[ERROR] No modules parsed from Modulplan.", file=sys.stderr)
        sys.exit(1)

    # Resolve any teaching-day deficit automatically by shortening the
    # Prüfungsvorbereitung (never the FPB). No interactive prompts — this keeps
    # the run fully unattended, also under Copilot agent mode.
    teaching_days = [d for d, dt in calendar_days if dt is None]
    available = len(teaching_days)
    deficit = check_deficit(teaching_days, modules, args.ue_pro_tag)

    if deficit < 0:
        total_needed = sum(math.ceil(m["ue"] / args.ue_pro_tag) for m in modules)
        print(
            f"[INFO] Defizit: {abs(deficit)} Unterrichtstag(e) "
            f"(benötigt: {total_needed}, verfügbar: {available}). "
            "Kürze automatisch die Prüfungsvorbereitung."
        )
        modules, adjustments = auto_adjust_pruefungsvorbereitung(
            modules, available, args.ue_pro_tag
        )
        if adjustments is None:
            print(
                f"[WARN] Kein Prüfungsvorbereitungs-Modul gefunden, das die "
                f"{abs(deficit)} fehlenden Tage aufnehmen kann. Plan wird ohne "
                "Kürzung erstellt (letzte Module evtl. unvollständig).",
                file=sys.stderr,
            )
        else:
            for adj in adjustments:
                print(
                    f"  -> {adj['id']}: {int(adj['old_ue'])} -> {int(adj['new_ue'])} UE "
                    f"({adj['days_removed']} Tag(e) gekürzt)."
                )
            new_deficit = check_deficit(teaching_days, modules, args.ue_pro_tag)
            if new_deficit < 0:
                print(
                    f"[WARN] Auch nach Kürzung der Prüfungsvorbereitung verbleibt "
                    f"ein Defizit von {abs(new_deficit)} Tag(en).",
                    file=sys.stderr,
                )
            else:
                print("[INFO] Defizit vollständig durch Kürzung der Prüfungsvorbereitung ausgeglichen.")
    else:
        print(f"[INFO] Kapazität: {deficit} freie(r) Unterrichtstag(e) nach Verplanung aller Module.")

    # Schedule
    schedule, day_to_module = schedule_modules(calendar_days, modules, args.ue_pro_tag)

    # Build Excel
    build_excel(schedule, calendar_days, day_to_module, args.kursname, output_path, args.ue_pro_tag)

    print("\n[DONE]")
    print(f"  Output:        {output_path}")
    print(f"  Teaching days: {len(teaching_days)}")
    print(f"  Modules:       {len(schedule)}")
    print(f"  Total UE:      {sum(m['ue'] for m in modules):.0f}")


if __name__ == "__main__":
    main()
